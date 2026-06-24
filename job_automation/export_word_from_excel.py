from __future__ import annotations

from datetime import datetime
from pathlib import Path

import typer
from openpyxl import load_workbook
from rich.console import Console

from config import EXPORT_DIR, MAX_COMPANY_EMPLOYEES
from exporters.docx_exporter import export_records_to_docx
from matching.company_intel import exceeds_employee_limit


app = typer.Typer(help="Generate a Word report from the edited Excel job workbook.")
console = Console()


def latest_excel_export() -> Path:
    files = sorted(EXPORT_DIR.glob("jobs_export_*_pretty.xlsx"), key=lambda path: path.stat().st_mtime, reverse=True)
    if not files:
        raise FileNotFoundError("No pretty Excel export found in data/exports.")
    return files[0]


def read_records_from_excel(path: Path, include_dismissed: bool) -> list[dict[str, str | int]]:
    workbook = load_workbook(path, data_only=True)
    if "Job Details" not in workbook.sheetnames:
        raise ValueError("Excel workbook must contain a 'Job Details' sheet.")

    sheet = workbook["Job Details"]
    headers = [str(cell.value or "").strip() for cell in sheet[1]]
    header_map = {header: index + 1 for index, header in enumerate(headers)}

    required_headers = {
        "Search Run Timestamp": "search_run_at",
        "Score": "score",
        "Priority": "priority",
        "Job Title": "title",
        "Company": "company",
        "Location": "location",
        "Remote / Hybrid / Onsite": "remote_type",
        "Potential Compensation": "salary",
        "Compensation >= 50k/year or >= 4,200 gross/month?": "salary_target_met",
        "Working Hours <= 36h/week?": "hours_target_met",
        "Compensation / Hours Notes": "salary_hour_notes",
        "Required Skills": "skills",
        "Requirements / Short Description": "description",
        "Job Posting Link": "url",
        "Why This Job Matches": "reason",
        "Dismissed?": "dismissed",
    }
    optional_headers = {
        "Company Type": "company_fit",
        "Company Size": "company_size",
        "Company Size Source": "company_size_source",
        "City": "city",
        "Country": "country",
        "Status": "status",
        "Interesting?": "interesting",
        "Next Step": "next_step",
        "Review Notes": "review_notes",
    }
    missing = [header for header in required_headers if header not in header_map]
    if missing:
        raise ValueError(f"Missing required Excel columns: {', '.join(missing)}")

    records: list[dict[str, str | int]] = []
    for row in range(2, sheet.max_row + 1):
        record: dict[str, str | int] = {}
        for header, key in required_headers.items():
            value = sheet.cell(row=row, column=header_map[header]).value
            record[key] = value if value is not None else ""
        for header, key in optional_headers.items():
            if header in header_map:
                value = sheet.cell(row=row, column=header_map[header]).value
                record[key] = value if value is not None else ""
        record.setdefault("source", "")
        record.setdefault("date_found", "")
        record.setdefault("date_posted", "")
        record.setdefault("status", "")
        record.setdefault("country", "")
        record.setdefault("region", "")
        record.setdefault("seniority", "")

        dismissed = str(record.get("dismissed", "")).strip().lower() == "yes"
        status_dismissed = str(record.get("status", "")).strip().lower() == "dismissed"
        if (dismissed or status_dismissed) and not include_dismissed:
            continue
        if exceeds_employee_limit(str(record.get("company_size", "")), MAX_COMPANY_EMPLOYEES):
            continue
        records.append(record)
    return records


def parse_bool(value: str | bool) -> bool:
    if isinstance(value, bool):
        return value
    return value.strip().lower() in {"1", "true", "yes", "y", "on"}


@app.command()
def main(
    excel: Path | None = typer.Option(None, help="Path to the edited Excel export. Defaults to latest pretty export."),
    include_dismissed: str = typer.Option("false", help="Include rows where Dismissed? is yes: true/false."),
    output: Path | None = typer.Option(None, help="Optional output .docx path."),
) -> None:
    excel_path = excel or latest_excel_export()
    records = read_records_from_excel(excel_path, include_dismissed=parse_bool(include_dismissed))
    output_path = output or (EXPORT_DIR / f"jobs_report_from_excel_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.docx")
    export_records_to_docx(records, output_path)
    console.print(f"Read {len(records)} jobs from {excel_path}")
    console.print(f"Word report saved to {output_path}")


if __name__ == "__main__":
    app()
