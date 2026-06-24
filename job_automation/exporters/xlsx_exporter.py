from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

from config import EXPORT_DIR
from database.models import Job
from exporters.job_presenter import build_job_records


XLSX_COLUMNS = [
    ("search_run_at", "Search Run Timestamp"),
    ("score", "Score"),
    ("title_fit_score", "Title Fit"),
    ("seniority_fit_score", "Seniority Fit"),
    ("remote_fit_score", "Remote Fit"),
    ("skill_fit_score", "Skill Fit"),
    ("geography_fit_score", "Geo Fit"),
    ("compensation_fit_score", "Comp / Hours Fit"),
    ("company_fit_score", "Company Signal Fit"),
    ("penalty_score", "Penalty"),
    ("priority", "Priority"),
    ("title", "Job Title"),
    ("company", "Company"),
    ("company_fit", "Company Type"),
    ("company_size", "Company Size"),
    ("company_size_source", "Company Size Source"),
    ("location", "Location"),
    ("city", "City"),
    ("country", "Country"),
    ("region", "Region"),
    ("remote_type", "Remote / Hybrid / Onsite"),
    ("seniority", "Level"),
    ("salary", "Potential Compensation"),
    ("salary_found", "Compensation Found?"),
    ("salary_target_met", "Compensation >= 50k/year or >= 4,200 gross/month?"),
    ("hours_found", "Working Hours Found?"),
    ("hours_target_met", "Working Hours <= 36h/week?"),
    ("salary_hour_notes", "Compensation / Hours Notes"),
    ("skills", "Required Skills"),
    ("description", "Requirements / Short Description"),
    ("source", "Source"),
    ("url", "Job Posting Link"),
    ("reason", "Why This Job Matches"),
    ("date_found", "Date Found"),
    ("date_posted", "Date Posted"),
    ("status", "Status"),
    ("interesting", "Interesting?"),
    ("applied", "Applied?"),
    ("next_step", "Next Step"),
    ("review_notes", "Review Notes"),
    ("dismissed", "Dismissed?"),
]


def export_jobs_to_xlsx(jobs: list[Job], export_dir: Path = EXPORT_DIR) -> Path:
    export_dir.mkdir(parents=True, exist_ok=True)
    path = export_dir / f"jobs_export_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}_pretty.xlsx"
    records = build_job_records(jobs)

    workbook = Workbook()
    overview = workbook.active
    overview.title = "Overview"
    details = workbook.create_sheet("Job Details")

    _build_overview(overview, records)
    _build_details(details, records)
    workbook.save(path)
    return path


def _build_overview(sheet, records: list[dict[str, str | int]]) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.merge_cells("A1:J1")
    sheet["A1"] = "AI Job Automation Export"
    sheet["A1"].font = Font(bold=True, size=20, color="111827")
    sheet["A1"].fill = PatternFill("solid", fgColor="D9EAF7")
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 40

    urgent = sum(1 for record in records if record["priority"] == "urgent")
    high = sum(1 for record in records if record["priority"] == "high")
    avg_score = round(sum(int(record["score"]) for record in records) / len(records), 1) if records else 0
    target_salary = sum(1 for record in records if record["salary_target_met"] == "yes")
    target_hours = sum(1 for record in records if record["hours_target_met"] == "yes")

    summary_rows = [
        ("Exported Jobs", len(records)),
        ("Average Score", avg_score),
        ("Urgent / High", f"{urgent} / {high}"),
        ("Compensation Target Met", target_salary),
        ("Hours Target Met", target_hours),
    ]
    for row_index, (label, value) in enumerate(summary_rows, start=3):
        label_cell = sheet.cell(row=row_index, column=1, value=label)
        value_cell = sheet.cell(row=row_index, column=2, value=value)
        label_cell.font = Font(bold=True)
        label_cell.fill = PatternFill("solid", fgColor="F3F4F6")
        value_cell.font = Font(bold=True)

    headers = ["Score", "Priority", "Status", "Job Title", "Company", "Company Type", "Company Size", "City", "Country", "Location", "Remote", "Compensation", "Skills", "Link", "Dismissed?"]
    sheet.append(headers)
    header_row = 9
    _style_header(sheet, header_row)

    for record in records:
        sheet.append(
            [
                record["score"],
                record["priority"],
                record["status"],
                record["title"],
                record["company"],
                record["company_fit"],
                record["company_size"],
                record["city"],
                record["country"],
                record["location"],
                record["remote_type"],
                record["salary"],
                record["skills"],
                record["url"],
                record["dismissed"],
            ]
        )

    sheet.freeze_panes = "A10"
    sheet.auto_filter.ref = f"A9:O{max(9, sheet.max_row)}"
    _set_widths(sheet, [9, 13, 16, 36, 24, 18, 18, 18, 18, 30, 16, 22, 42, 58, 14])
    _wrap_range(sheet, 10, max(10, sheet.max_row), 3, 15)
    _apply_score_formatting(sheet, f"A10:A{max(10, sheet.max_row)}")
    _apply_review_workflow(sheet, 10, max(10, sheet.max_row), 3, 15, 15)
    _apply_table_borders(sheet, 3, 1, max(sheet.max_row, 9), 15)


def _build_details(sheet, records: list[dict[str, str | int]]) -> None:
    sheet.sheet_view.showGridLines = False
    headers = [label for _, label in XLSX_COLUMNS]
    sheet.append(headers)
    _style_header(sheet, 1)

    for record in records:
        sheet.append([record.get(key, "") for key, _ in XLSX_COLUMNS])

    end_row = max(2, sheet.max_row)
    table_ref = f"A1:{get_column_letter(len(XLSX_COLUMNS))}{end_row}"
    table = Table(displayName="JobsTable", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    sheet.add_table(table)
    sheet.freeze_panes = "A2"

    _set_widths(sheet, [22, 9, 10, 12, 10, 10, 10, 16, 18, 10, 13, 34, 22, 18, 18, 22, 30, 18, 14, 14, 22, 16, 22, 18, 30, 18, 22, 36, 44, 70, 16, 58, 52, 14, 20, 14, 16, 12, 22, 36, 14])
    _wrap_range(sheet, 2, end_row, 3, len(XLSX_COLUMNS))
    _apply_score_formatting(sheet, f"B2:B{end_row}")
    salary_found_col = _column_index("salary_found")
    hours_target_col = _column_index("hours_target_met")
    _apply_yes_no_formatting(sheet, f"{get_column_letter(salary_found_col)}2:{get_column_letter(hours_target_col)}{end_row}")
    _apply_yes_no_validation(sheet, 2, end_row, len(XLSX_COLUMNS) - 4)
    _apply_yes_no_validation(sheet, 2, end_row, len(XLSX_COLUMNS) - 3)
    _apply_review_workflow(sheet, 2, end_row, len(XLSX_COLUMNS) - 5, len(XLSX_COLUMNS), len(XLSX_COLUMNS))
    for row in range(2, end_row + 1):
        sheet.row_dimensions[row].height = 72


def _style_header(sheet, row: int) -> None:
    for cell in sheet[row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _set_widths(sheet, widths: list[int]) -> None:
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def _column_index(key: str) -> int:
    return [column_key for column_key, _ in XLSX_COLUMNS].index(key) + 1


def _wrap_range(sheet, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            sheet.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="top")


def _apply_score_formatting(sheet, cell_range: str) -> None:
    sheet.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="greaterThanOrEqual", formula=["80"], fill=PatternFill("solid", fgColor="C6EFCE")),
    )
    sheet.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="between", formula=["60", "79"], fill=PatternFill("solid", fgColor="FFEB9C")),
    )


def _apply_yes_no_formatting(sheet, cell_range: str) -> None:
    sheet.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="equal", formula=['"yes"'], fill=PatternFill("solid", fgColor="C6EFCE")),
    )
    sheet.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="equal", formula=['"no"'], fill=PatternFill("solid", fgColor="FFC7CE")),
    )


def _apply_yes_no_validation(sheet, start_row: int, end_row: int, column: int) -> None:
    letter = get_column_letter(column)
    validation = DataValidation(type="list", formula1='"no,yes"', allow_blank=True)
    sheet.add_data_validation(validation)
    validation.add(f"{letter}{start_row}:{letter}{end_row}")


def _apply_review_workflow(sheet, start_row: int, end_row: int, status_col: int, end_col: int, dismissed_col: int) -> None:
    status_letter = get_column_letter(status_col)
    status_validation = DataValidation(type="list", formula1='"new,shortlist,applied,interview,dismissed"', allow_blank=True)
    sheet.add_data_validation(status_validation)
    status_validation.add(f"{status_letter}{start_row}:{status_letter}{end_row}")

    dismissed_letter = get_column_letter(dismissed_col)
    validation = DataValidation(type="list", formula1='"no,yes"', allow_blank=False)
    sheet.add_data_validation(validation)
    validation.add(f"{dismissed_letter}{start_row}:{dismissed_letter}{end_row}")
    row_range = f"A{start_row}:{get_column_letter(end_col)}{end_row}"
    sheet.conditional_formatting.add(
        row_range,
        FormulaRule(
            formula=[f'OR(${dismissed_letter}{start_row}="yes",${status_letter}{start_row}="dismissed")'],
            font=Font(strike=True, color="808080"),
            fill=PatternFill("solid", fgColor="F2F2F2"),
        ),
    )


def _apply_table_borders(sheet, start_row: int, start_col: int, end_row: int, end_col: int) -> None:
    side = Side(style="thin", color="D9E2EC")
    border = Border(left=side, right=side, top=side, bottom=side)
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            sheet.cell(row=row, column=col).border = border
