from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from database.models import Job
from export_word_from_excel import read_records_from_excel
from exporters.xlsx_exporter import export_jobs_to_xlsx
from main import score_jobs
from matching.company_intel import exceeds_employee_limit, parse_employee_floor
from matching.deduplication import deduplicate_jobs
from matching.region_detection import is_location_name
from scrapers.generic_search_scraper import GenericSearchScraper
from scrapers.mock_website_scraper import MockWebsiteScraper

from config import MAX_COMPANY_EMPLOYEES


@pytest.mark.parametrize(
    "company_size, expected",
    [
        ("unknown", False),
        ("", False),
        ("1-10", False),
        ("11-50", False),
        ("51-200", False),  # upper bound is exactly the limit -> kept
        ("51-500", False),  # weak SME inference -> kept (floor 51)
        ("201-500", True),
        ("1,001-5,000", True),
        ("1,001+", True),
        ("200+", True),  # open-ended "at least 200" -> dropped
        ("10,001+", True),
    ],
)
def test_exceeds_employee_limit(company_size: str, expected: bool):
    assert exceeds_employee_limit(company_size, 200) is expected


def test_parse_employee_floor_uses_lower_bound():
    assert parse_employee_floor("51-200") == 51
    assert parse_employee_floor("1,001+") == 1001
    assert parse_employee_floor("unknown") is None
    assert parse_employee_floor("") is None


def test_is_location_name_rejects_places_but_keeps_real_companies():
    assert is_location_name("Egypt") is True
    assert is_location_name("Cairo") is True
    assert is_location_name("Cairo, Egypt") is True
    assert is_location_name("Remote") is True
    # Real company names that merely contain a place word must survive.
    assert is_location_name("London Fintech Ltd") is False
    assert is_location_name("FlowPilot AI") is False
    assert is_location_name("Berlin Brands Group") is False


def test_generic_scraper_does_not_use_a_country_as_company():
    scraper = GenericSearchScraper(limit=10)
    # Geographic name reached through the "at <place>" pattern.
    assert scraper._guess_company("AI Automation Specialist at Egypt", "tanqeeb.com") == "Unknown"
    # Geographic name reached through the trailing title piece.
    assert scraper._guess_company("AI Automation Specialist - Cairo, Egypt", "tanqeeb.com") == "Unknown"
    # A genuine company is still recognised.
    assert scraper._guess_company("Data Scientist at Acme Inc", "greenhouse.io") == "Acme Inc"


def test_known_large_companies_are_dropped_via_size_override():
    # Well-known giants (HP, IBM, Orange Business...) are blacklisted in
    # company_sizes.yaml so the <200 filter bites even when the posting never
    # states a headcount. They are dropped; a normal-named company is kept.
    from matching.company_intel import lookup_company_size

    for giant in ("HP", "IBM", "Orange Business", "Hewlett Packard Enterprise"):
        override = lookup_company_size(giant)
        assert override is not None, f"{giant} should be in the size blacklist"
        assert exceeds_employee_limit(override[0], MAX_COMPANY_EMPLOYEES) is True

    assert lookup_company_size("FlowPilot AI") is None


def test_mock_pipeline_drops_companies_over_200_employees():
    jobs = MockWebsiteScraper(limit=10).search(region="europe", remote=True)
    scored = sorted(score_jobs(deduplicate_jobs(jobs)), key=lambda job: job.relevance_score, reverse=True)

    big = next(job for job in scored if job.company_name == "ScaleCorp Enterprise")
    # The role itself is a strong match, so only the size filter should remove it.
    assert big.relevance_score >= 70
    assert exceeds_employee_limit(big.company_size, MAX_COMPANY_EMPLOYEES) is True

    kept = [job for job in scored if not exceeds_employee_limit(job.company_size, MAX_COMPANY_EMPLOYEES)]
    kept_companies = {job.company_name for job in kept}
    assert "ScaleCorp Enterprise" not in kept_companies
    assert "FlowPilot AI" in kept_companies


def test_word_from_excel_excludes_large_companies(tmp_path: Path):
    jobs = [
        Job(
            job_title="Junior AI Automation Specialist",
            company_name="Small Startup",
            location="Remote Germany",
            remote_type="remote",
            job_url="https://example.com/small",
            company_size="11-50",
            relevance_score=90,
            priority_level="urgent",
            reason_for_score="Great fit.",
        ),
        Job(
            job_title="AI Automation Engineer",
            company_name="Big Corp",
            location="Remote Europe",
            remote_type="remote",
            job_url="https://example.com/big",
            company_size="1,001-5,000",
            relevance_score=85,
            priority_level="urgent",
            reason_for_score="Good fit but large company.",
        ),
    ]
    excel_path = export_jobs_to_xlsx(jobs, tmp_path)

    # Sanity check: both rows are written to the workbook.
    workbook = load_workbook(excel_path)
    assert workbook["Job Details"].max_row == 3

    records = read_records_from_excel(excel_path, include_dismissed=True)
    companies = {record["company"] for record in records}
    assert "Small Startup" in companies
    assert "Big Corp" not in companies
